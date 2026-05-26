import os, sqlite3, secrets
try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None
from datetime import datetime
from functools import wraps
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, abort
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR=os.path.abspath(os.path.dirname(__file__))
DB_PATH=os.path.join(BASE_DIR,'dental.db')
app=Flask(__name__)
app.secret_key=os.environ.get('SECRET_KEY','dev-secret-change-me')

GRADES=[f'ม.{i}' for i in range(1,7)]
GROWTH_WA=['น้ำหนักน้อยกว่าเกณฑ์','น้ำหนักตามเกณฑ์','น้ำหนักมากกว่าเกณฑ์']
GROWTH_HA=['เตี้ย','ค่อนข้างเตี้ย','ส่วนสูงตามเกณฑ์','ค่อนข้างสูง','สูง']
GROWTH_WH=['ผอม','ค่อนข้างผอม','สมส่วน','ท้วม','เริ่มอ้วน','อ้วน']
NUTRITIONS=GROWTH_WH

def now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

THAI_MONTHS = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']

def thai_date(value):
    if not value:
        return ''
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
            try:
                value = datetime.strptime(value[:19] if fmt.endswith('%S') else value[:10], fmt)
                break
            except Exception:
                pass
        else:
            return value
    return f'{value.day} {THAI_MONTHS[value.month]} {value.year + 543}'

def buddhist_year():
    return datetime.now().year + 543

@app.context_processor
def utility_processor():
    return dict(thai_date=thai_date, buddhist_year=buddhist_year)


def next_round_title(con):
    year = buddhist_year()
    n = con.execute('SELECT COUNT(*) AS c FROM rounds WHERE title LIKE ?', (f'%/{year}',)).fetchone()['c'] + 1
    return f'รอบตรวจสุขภาพครั้งที่ {n}/{year}'
def is_postgres():
    return bool(os.environ.get('DATABASE_URL'))

class PgConnection:
    def __init__(self):
        if psycopg2 is None:
            raise RuntimeError('DATABASE_URL is set but psycopg2-binary is not installed')
        url=os.environ.get('DATABASE_URL')
        if url and url.startswith('postgres://'):
            url='postgresql://' + url[len('postgres://'):]
        self.con=psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
    def execute(self, sql, params=()):
        sql = sql.replace('?', '%s')
        cur=self.con.cursor()
        cur.execute(sql, params)
        return cur
    def commit(self):
        self.con.commit()
    def close(self):
        self.con.close()

def db():
    if is_postgres():
        return PgConnection()
    con=sqlite3.connect(DB_PATH)
    con.row_factory=sqlite3.Row
    return con

def init_db():
    con=db()
    if is_postgres():
        con.execute("""
        CREATE TABLE IF NOT EXISTS users(
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT CHECK(role IN ('admin','user')) NOT NULL,
            created_at TEXT
        )""")
        con.execute("""
        CREATE TABLE IF NOT EXISTS rounds(
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            school TEXT,
            school_address TEXT,
            village_no TEXT,
            subdistrict TEXT,
            district TEXT,
            province TEXT,
            zipcode TEXT,
            phone TEXT,
            survey_date TEXT,
            is_open INTEGER DEFAULT 1,
            public_token TEXT UNIQUE,
            created_at TEXT,
            updated_at TEXT
        )""")
        con.execute("""
        CREATE TABLE IF NOT EXISTS students(
            id SERIAL PRIMARY KEY,
            round_id INTEGER REFERENCES rounds(id),
            id_card TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            birthdate TEXT,
            gender TEXT,
            grade TEXT,
            room TEXT,
            address TEXT,
            weight REAL,
            height REAL,
            waist REAL,
            hip REAL,
            growth_weight_age TEXT,
            growth_height_age TEXT,
            growth_weight_height TEXT,
            nutrition TEXT,
            tooth_decay INTEGER DEFAULT 0,
            gum_disease INTEGER DEFAULT 0,
            urgent INTEGER DEFAULT 0,
            note TEXT,
            created_at TEXT,
            updated_at TEXT,
            created_by TEXT,
            updated_by TEXT
        )""")
        con.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs(
            id SERIAL PRIMARY KEY,
            actor TEXT,
            action TEXT,
            detail TEXT,
            created_at TEXT
        )""")
        if not con.execute('SELECT id FROM users WHERE username=%s',('admin',)).fetchone():
            con.execute('INSERT INTO users(username,password_hash,role,created_at) VALUES(%s,%s,%s,%s)',('admin',generate_password_hash('admin123'),'admin',now()))
        if not con.execute('SELECT id FROM users WHERE username=%s',('user',)).fetchone():
            con.execute('INSERT INTO users(username,password_hash,role,created_at) VALUES(%s,%s,%s,%s)',('user',generate_password_hash('user123'),'user',now()))
        if not con.execute('SELECT id FROM rounds LIMIT 1').fetchone():
            con.execute('INSERT INTO rounds(title,school,school_address,village_no,subdistrict,district,province,zipcode,phone,survey_date,is_open,public_token,created_at,updated_at) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(f'รอบตรวจสุขภาพครั้งที่ 1/{buddhist_year()}','','','','','','ขอนแก่น','','',datetime.now().strftime('%Y-%m-%d'),1,secrets.token_urlsafe(10),now(),now()))
        con.commit(); con.close(); return

    cur=con.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password_hash TEXT, role TEXT CHECK(role IN ('admin','user')) NOT NULL, created_at TEXT);
    CREATE TABLE IF NOT EXISTS rounds(id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL, school TEXT, school_address TEXT, village_no TEXT, subdistrict TEXT, district TEXT, province TEXT, zipcode TEXT, phone TEXT, survey_date TEXT, is_open INTEGER DEFAULT 1, public_token TEXT UNIQUE, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS students(id INTEGER PRIMARY KEY AUTOINCREMENT, round_id INTEGER, id_card TEXT NOT NULL UNIQUE, name TEXT NOT NULL, birthdate TEXT, gender TEXT, grade TEXT, room TEXT, address TEXT, weight REAL, height REAL, waist REAL, hip REAL, growth_weight_age TEXT, growth_height_age TEXT, growth_weight_height TEXT, nutrition TEXT, tooth_decay INTEGER DEFAULT 0, gum_disease INTEGER DEFAULT 0, urgent INTEGER DEFAULT 0, note TEXT, created_at TEXT, updated_at TEXT, created_by TEXT, updated_by TEXT, FOREIGN KEY(round_id) REFERENCES rounds(id));
    CREATE TABLE IF NOT EXISTS audit_logs(id INTEGER PRIMARY KEY AUTOINCREMENT, actor TEXT, action TEXT, detail TEXT, created_at TEXT);
    """)
    if not cur.execute('SELECT id FROM users WHERE username=?',('admin',)).fetchone():
        cur.execute('INSERT INTO users(username,password_hash,role,created_at) VALUES(?,?,?,?)',('admin',generate_password_hash('admin123'),'admin',now()))
    if not cur.execute('SELECT id FROM users WHERE username=?',('user',)).fetchone():
        cur.execute('INSERT INTO users(username,password_hash,role,created_at) VALUES(?,?,?,?)',('user',generate_password_hash('user123'),'user',now()))
    if not cur.execute('SELECT id FROM rounds').fetchone():
        cur.execute('INSERT INTO rounds(title,school,school_address,village_no,subdistrict,district,province,zipcode,phone,survey_date,is_open,public_token,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(f'รอบตรวจสุขภาพครั้งที่ 1/{buddhist_year()}','','','','','','ขอนแก่น','','',datetime.now().strftime('%Y-%m-%d'),1,secrets.token_urlsafe(10),now(),now()))
    rcols=[r[1] for r in cur.execute('PRAGMA table_info(rounds)').fetchall()]
    for name,typ in [('school_address','TEXT'),('village_no','TEXT'),('subdistrict','TEXT'),('district','TEXT'),('province','TEXT'),('zipcode','TEXT'),('phone','TEXT')]:
        if name not in rcols:
            cur.execute(f'ALTER TABLE rounds ADD COLUMN {name} {typ}')
    cols=[r[1] for r in cur.execute('PRAGMA table_info(students)').fetchall()]
    for name,typ in [('hip','REAL'),('growth_weight_age','TEXT'),('growth_height_age','TEXT'),('growth_weight_height','TEXT')]:
        if name not in cols:
            cur.execute(f'ALTER TABLE students ADD COLUMN {name} {typ}')
    con.commit(); con.close()

def log(action,detail=''):
    con=db(); con.execute('INSERT INTO audit_logs(actor,action,detail,created_at) VALUES(?,?,?,?)',(session.get('username','public'),action,detail,now())); con.commit(); con.close()

def current_user(): return session.get('username')
def login_required(f):
    @wraps(f)
    def w(*a,**k):
        if not session.get('user_id'): return redirect(url_for('login'))
        return f(*a,**k)
    return w
def admin_required(f):
    @wraps(f)
    def w(*a,**k):
        if session.get('role')!='admin': abort(403)
        return f(*a,**k)
    return w

def stats(round_id=None):
    con=db(); where=''; params=[]
    if round_id: where='WHERE round_id=?'; params=[round_id]
    rows=con.execute(f'SELECT * FROM students {where}',params).fetchall()
    total=len(rows)
    by_grade={g:{'total':0,'decay':0,'gum':0,'urgent':0} for g in GRADES}
    nut={n:0 for n in NUTRITIONS}
    for r in rows:
        g=r['grade'] or 'ไม่ระบุ'
        by_grade.setdefault(g,{'total':0,'decay':0,'gum':0,'urgent':0})
        by_grade[g]['total']+=1; by_grade[g]['decay']+=r['tooth_decay']; by_grade[g]['gum']+=r['gum_disease']; by_grade[g]['urgent']+=r['urgent']
        label=r['growth_weight_height'] or r['nutrition']
        if label in nut: nut[label]+=1
    return {'total':total,'decay':sum(r['tooth_decay'] for r in rows),'gum':sum(r['gum_disease'] for r in rows),'urgent':sum(r['urgent'] for r in rows),'by_grade':by_grade,'nutrition':nut}

@app.route('/init-db')
def init_db_route():
    init_db()
    return 'Database initialized successfully'

@app.route('/')
def home(): return redirect(url_for('dashboard'))
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        con=db(); u=con.execute('SELECT * FROM users WHERE username=?',(request.form['username'],)).fetchone(); con.close()
        if u and check_password_hash(u['password_hash'],request.form['password']):
            session.update(user_id=u['id'],username=u['username'],role=u['role']); return redirect(url_for('dashboard'))
        flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง','danger')
    return render_template('login.html')
@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    con=db(); rounds=con.execute('SELECT * FROM rounds ORDER BY id DESC').fetchall(); rid=request.args.get('round_id', type=int)
    if not rid and rounds: rid=rounds[0]['id']
    s=stats(rid); con.close(); return render_template('dashboard.html',stats=s,rounds=rounds,round_id=rid,now_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/rounds',methods=['GET','POST'])
@login_required
def rounds():
    con=db()
    if request.method=='POST':
        title = request.form.get('title') or next_round_title(con)
        survey_date = request.form.get('survey_date') or datetime.now().strftime('%Y-%m-%d')
        cur=con.execute('''INSERT INTO rounds(title,school,school_address,village_no,subdistrict,district,province,zipcode,phone,survey_date,is_open,public_token,created_at,updated_at)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                        (title,request.form.get('school',''),request.form.get('school_address',''),request.form.get('village_no',''),request.form.get('subdistrict',''),request.form.get('district',''),request.form.get('province',''),request.form.get('zipcode',''),request.form.get('phone',''),survey_date,1,secrets.token_urlsafe(10),now(),now()))
        con.commit(); log('create_round',title); return redirect(url_for('rounds'))
    rows=con.execute('SELECT * FROM rounds ORDER BY id DESC').fetchall(); con.close(); return render_template('rounds.html',rounds=rows)

@app.route('/round/<int:rid>/edit',methods=['GET','POST'])
@login_required
def edit_round(rid):
    con=db()
    r=con.execute('SELECT * FROM rounds WHERE id=?',(rid,)).fetchone()
    if not r:
        con.close(); abort(404)
    if request.method=='POST':
        con.execute('''UPDATE rounds SET title=?, school=?, school_address=?, village_no=?, subdistrict=?, district=?, province=?, zipcode=?, phone=?, survey_date=?, is_open=?, updated_at=? WHERE id=?''',
                    (request.form['title'], request.form.get('school',''), request.form.get('school_address',''), request.form.get('village_no',''), request.form.get('subdistrict',''), request.form.get('district',''), request.form.get('province',''), request.form.get('zipcode',''), request.form.get('phone',''), request.form.get('survey_date',''), 1 if request.form.get('is_open') else 0, now(), rid))
        con.commit(); con.close(); log('edit_round',str(rid)); return redirect(url_for('rounds'))
    con.close(); return render_template('round_form.html',r=r)

@app.route('/round/<int:rid>/toggle')
@login_required
def toggle_round(rid):
    con=db(); con.execute('UPDATE rounds SET is_open=1-is_open, updated_at=? WHERE id=?',(now(),rid)); con.commit(); con.close(); return redirect(url_for('rounds'))
@app.route('/round/<int:rid>/delete',methods=['POST'])
@login_required
def delete_round(rid):
    con=db(); con.execute('DELETE FROM students WHERE round_id=?',(rid,)); con.execute('DELETE FROM rounds WHERE id=?',(rid,)); con.commit(); con.close(); log('delete_round',str(rid)); return redirect(url_for('rounds'))

@app.route('/students')
@login_required
def students():
    q=request.args.get('q','').strip()
    rid=request.args.get('round_id', type=int)
    grade=request.args.get('grade','').strip()
    room=request.args.get('room','').strip()
    con=db()
    rounds=con.execute('SELECT * FROM rounds ORDER BY id DESC').fetchall()
    room_rows=con.execute("SELECT DISTINCT room FROM students WHERE room IS NOT NULL AND TRIM(room)<>'' ORDER BY room").fetchall()
    rooms=[x['room'] for x in room_rows]
    sql='SELECT s.*, r.title round_title FROM students s LEFT JOIN rounds r ON r.id=s.round_id WHERE 1=1'
    params=[]
    if rid:
        sql+=' AND s.round_id=?'
        params.append(rid)
    if q:
        sql+=' AND (s.id_card LIKE ? OR s.name LIKE ?)'
        params += [f'%{q}%', f'%{q}%']
    if grade:
        sql+=' AND s.grade=?'
        params.append(grade)
    if room:
        sql+=' AND s.room=?'
        params.append(room)
    rows=con.execute(sql+' ORDER BY s.grade, s.room, s.name, s.id DESC',params).fetchall()
    con.close()
    return render_template('students.html',students=rows,rounds=rounds,q=q,round_id=rid,grade=grade,room=room,grades=GRADES,rooms=rooms)
@app.route('/student/add',methods=['GET','POST'])
@login_required
def add_student(): return save_student(None)
@app.route('/student/<int:sid>/edit',methods=['GET','POST'])
@login_required
def edit_student(sid): return save_student(sid)

def save_student(sid):
    con=db(); rounds=con.execute('SELECT * FROM rounds ORDER BY id DESC').fetchall(); student=None
    if sid: student=con.execute('SELECT * FROM students WHERE id=?',(sid,)).fetchone()
    if request.method=='POST':
        data={k:request.form.get(k,'').strip() for k in ['id_card','name','birthdate','gender','grade','room','address','growth_weight_age','growth_height_age','growth_weight_height','nutrition','note']}
        rid=int(request.form.get('round_id') or rounds[0]['id'])
        weight=request.form.get('weight') or None; height=request.form.get('height') or None; waist=request.form.get('waist') or None; hip=request.form.get('hip') or None
        decay=1 if request.form.get('tooth_decay') else 0; gum=1 if request.form.get('gum_disease') else 0; urgent=1 if request.form.get('urgent') else 0
        dup=con.execute('SELECT id FROM students WHERE id_card=? AND id<>?',(data['id_card'],sid or 0)).fetchone()
        if dup: flash('เลขบัตรประชาชนนี้ถูกกรอกแล้ว ระบบไม่ให้ซ้ำทั้งระบบ','danger'); return render_template('student_form.html',student=student,rounds=rounds,nutritions=NUTRITIONS,growth_wa=GROWTH_WA,growth_ha=GROWTH_HA,growth_wh=GROWTH_WH)
        if sid:
            con.execute('''UPDATE students SET round_id=?, id_card=?, name=?, birthdate=?, gender=?, grade=?, room=?, address=?, weight=?, height=?, waist=?, hip=?, growth_weight_age=?, growth_height_age=?, growth_weight_height=?, nutrition=?, tooth_decay=?, gum_disease=?, urgent=?, note=?, updated_at=?, updated_by=? WHERE id=?''',(rid,data['id_card'],data['name'],data['birthdate'],data['gender'],data['grade'],data['room'],data['address'],weight,height,waist,hip,data['growth_weight_age'],data['growth_height_age'],data['growth_weight_height'],data['growth_weight_height'],decay,gum,urgent,data['note'],now(),current_user(),sid)); log('edit_student',data['id_card'])
        else:
            con.execute('''INSERT INTO students(round_id,id_card,name,birthdate,gender,grade,room,address,weight,height,waist,hip,growth_weight_age,growth_height_age,growth_weight_height,nutrition,tooth_decay,gum_disease,urgent,note,created_at,updated_at,created_by,updated_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(rid,data['id_card'],data['name'],data['birthdate'],data['gender'],data['grade'],data['room'],data['address'],weight,height,waist,hip,data['growth_weight_age'],data['growth_height_age'],data['growth_weight_height'],data['growth_weight_height'],decay,gum,urgent,data['note'],now(),now(),current_user(),current_user())); log('add_student',data['id_card'])
        con.commit(); con.close(); return redirect(url_for('students'))
    con.close(); return render_template('student_form.html',student=student,rounds=rounds,nutritions=NUTRITIONS,growth_wa=GROWTH_WA,growth_ha=GROWTH_HA,growth_wh=GROWTH_WH)
@app.route('/student/<int:sid>/delete',methods=['POST'])
@login_required
def delete_student(sid):
    con=db(); con.execute('DELETE FROM students WHERE id=?',(sid,)); con.commit(); con.close(); log('delete_student',str(sid)); return redirect(url_for('students'))

@app.route('/form/<token>',methods=['GET','POST'])
def public_form(token):
    con=db(); r=con.execute('SELECT * FROM rounds WHERE public_token=?',(token,)).fetchone()
    if not r or not r['is_open']: abort(404)
    if request.method=='POST':
        idc=request.form.get('id_card','').strip(); name=request.form.get('name','').strip()
        if con.execute('SELECT id FROM students WHERE id_card=?',(idc,)).fetchone():
            flash('เลขบัตรประชาชนนี้เคยกรอกแล้ว ติดต่อเจ้าหน้าที่หากต้องแก้ไข','danger')
        else:
            con.execute('''INSERT INTO students(round_id,id_card,name,birthdate,gender,grade,room,address,weight,height,waist,hip,growth_weight_age,growth_height_age,growth_weight_height,nutrition,tooth_decay,gum_disease,urgent,note,created_at,updated_at,created_by,updated_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(r['id'],idc,name,request.form.get('birthdate',''),request.form.get('gender',''),request.form.get('grade',''),request.form.get('room',''),request.form.get('address',''),request.form.get('weight') or None,request.form.get('height') or None,request.form.get('waist') or None,request.form.get('hip') or None,request.form.get('growth_weight_age',''),request.form.get('growth_height_age',''),request.form.get('growth_weight_height',''),request.form.get('growth_weight_height',''),0,0,0,'',now(),now(),'public','public'))
            con.commit(); log('public_submit',idc); con.close(); return render_template('public_done.html')
    con.close(); return render_template('student_form.html',student=None,rounds=[r],nutritions=NUTRITIONS,growth_wa=GROWTH_WA,growth_ha=GROWTH_HA,growth_wh=GROWTH_WH,public=True,round_fixed=r)


@app.route('/dental-report')
def dental_report():
    con=db(); rounds=con.execute('SELECT * FROM rounds ORDER BY id DESC').fetchall(); rid=request.args.get('round_id', type=int)
    if not rid and rounds: rid=rounds[0]['id']
    r=con.execute('SELECT * FROM rounds WHERE id=?',(rid,)).fetchone() if rid else None
    rows=con.execute('SELECT * FROM students WHERE (? IS NULL OR round_id=?)',(rid,rid)).fetchall(); con.close()
    grades=GRADES
    total=len(rows)
    def c(grade, field, value):
        return sum(1 for x in rows if (x['grade'] or '')==grade and int(x[field] or 0)==value)
    report_rows=[]
    for problem, field, statuses in [
        ('ฟันผุ', 'tooth_decay', [('ไม่มี',0),('มี',1)]),
        ('เหงือกอักเสบ', 'gum_disease', [('ไม่มี',0),('มี',1)]),
        ('เร่งด่วน', 'urgent', [('เร่งด่วน',1)]),
    ]:
        for i,(status,val) in enumerate(statuses):
            counts={g:c(g,field,val) for g in grades}
            t=sum(counts.values())
            report_rows.append({'problem':problem,'status':status,'counts':counts,'total':t,'percent':(t*100/total if total else 0),'first':i==0,'rowspan':len(statuses)})
    decay=sum(int(x['tooth_decay'] or 0) for x in rows); gum=sum(int(x['gum_disease'] or 0) for x in rows); urgent=sum(int(x['urgent'] or 0) for x in rows)
    return render_template('dental_report.html',rounds=rounds,round_id=rid,round=r,grades=grades,total=total,report_rows=report_rows,decay_pct=(decay*100/total if total else 0),gum_pct=(gum*100/total if total else 0),urgent_pct=(urgent*100/total if total else 0))

@app.route('/export')
@login_required
def export_excel():
    rid=request.args.get('round_id',type=int)
    con=db()
    rows=con.execute("""SELECT s.*,r.title round_title FROM students s LEFT JOIN rounds r ON r.id=s.round_id
                        WHERE (? IS NULL OR s.round_id=?) ORDER BY s.grade, s.room, s.name, s.id""",(rid,rid)).fetchall()
    round_row=con.execute('SELECT * FROM rounds WHERE id=?',(rid,)).fetchone() if rid else None
    con.close()

    wb=Workbook()
    ws=wb.active
    ws.title='ข้อมูลนักเรียน'
    summary=wb.create_sheet('สรุปแดชบอร์ด')
    chart_sheet=wb.create_sheet('กราฟรายงาน')

    header_fill=PatternFill('solid', fgColor='D9EAF7')
    title_fill=PatternFill('solid', fgColor='1565C0')
    bold=Font(bold=True)
    thin=Side(style='thin', color='999999')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.append(['รอบ','เลขบัตร','ชื่อ-สกุล','เพศ','ชั้น','ห้อง','วันเกิด','ที่อยู่','น้ำหนัก','ส่วนสูง','รอบเอว','สะโพก','นน.อายุ','สส.อายุ','นน.สส.','ฟันผุ','เหงือกอักเสบ','เร่งด่วน','หมายเหตุ','สร้างเมื่อ','แก้ไขเมื่อ'])
    for cell in ws[1]:
        cell.fill=header_fill; cell.font=bold; cell.border=border; cell.alignment=Alignment(horizontal='center')
    for x in rows:
        ws.append([x['round_title'],x['id_card'],x['name'],x['gender'],x['grade'],x['room'],x['birthdate'],x['address'],x['weight'],x['height'],x['waist'],x['hip'],x['growth_weight_age'],x['growth_height_age'],x['growth_weight_height'],'มี' if x['tooth_decay'] else 'ไม่มี','มี' if x['gum_disease'] else 'ไม่มี','เร่งด่วน' if x['urgent'] else 'ไม่เร่งด่วน',x['note'],x['created_at'],x['updated_at']])
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.border=border
    widths=[22,18,28,10,10,10,14,32,10,10,10,10,18,18,18,12,14,14,24,18,18]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'

    total=len(rows)
    by_grade={g:{'total':0,'decay':0,'gum':0,'urgent':0} for g in GRADES}
    nutrition={n:0 for n in NUTRITIONS}
    for x in rows:
        g=x['grade'] or 'ไม่ระบุ'
        if g not in by_grade: by_grade[g]={'total':0,'decay':0,'gum':0,'urgent':0}
        by_grade[g]['total'] += 1
        by_grade[g]['decay'] += int(x['tooth_decay'] or 0)
        by_grade[g]['gum'] += int(x['gum_disease'] or 0)
        by_grade[g]['urgent'] += int(x['urgent'] or 0)
        label=x['growth_weight_height'] or x['nutrition']
        if label in nutrition: nutrition[label]+=1

    summary.merge_cells('A1:H1')
    summary['A1']='สรุปแดชบอร์ดรายงานผลภาวะช่องปากและโภชนาการ'
    summary['A1'].fill=title_fill; summary['A1'].font=Font(color='FFFFFF', bold=True, size=16); summary['A1'].alignment=Alignment(horizontal='center')
    summary['A2']='รอบ'; summary['B2']=round_row['title'] if round_row else 'ทั้งหมด'
    summary['A3']='วันที่ออกรายงาน'; summary['B3']=thai_date(datetime.now())
    summary['A4']='จำนวนนักเรียนทั้งหมด'; summary['B4']=total
    summary['D2']='ฟันผุ'; summary['E2']=sum(v['decay'] for v in by_grade.values())
    summary['D3']='เหงือกอักเสบ'; summary['E3']=sum(v['gum'] for v in by_grade.values())
    summary['D4']='เร่งด่วน'; summary['E4']=sum(v['urgent'] for v in by_grade.values())
    for c in ['A2','A3','A4','D2','D3','D4']:
        summary[c].font=bold; summary[c].fill=header_fill

    start_row=6
    headers=['ชั้น','ทั้งหมด','ฟันผุ','% ฟันผุ','เหงือกอักเสบ','% เหงือก','เร่งด่วน','% เร่งด่วน']
    for col,h in enumerate(headers,1):
        cell=summary.cell(start_row,col,h); cell.fill=header_fill; cell.font=bold; cell.border=border; cell.alignment=Alignment(horizontal='center')
    r=start_row+1
    for g,v in by_grade.items():
        summary.cell(r,1,g)
        summary.cell(r,2,v['total'])
        summary.cell(r,3,v['decay'])
        summary.cell(r,4,(v['decay']*100/v['total']) if v['total'] else 0)
        summary.cell(r,5,v['gum'])
        summary.cell(r,6,(v['gum']*100/v['total']) if v['total'] else 0)
        summary.cell(r,7,v['urgent'])
        summary.cell(r,8,(v['urgent']*100/v['total']) if v['total'] else 0)
        for col in range(1,9):
            summary.cell(r,col).border=border
            if col in [4,6,8]: summary.cell(r,col).number_format='0.0'
        r+=1

    nut_start=r+2
    summary.cell(nut_start,1,'ภาวะโภชนาการ').fill=header_fill; summary.cell(nut_start,1).font=bold
    summary.cell(nut_start,2,'จำนวน').fill=header_fill; summary.cell(nut_start,2).font=bold
    rr=nut_start+1
    for label,count in nutrition.items():
        summary.cell(rr,1,label); summary.cell(rr,2,count)
        summary.cell(rr,1).border=border; summary.cell(rr,2).border=border
        rr+=1
    for col in range(1,9): summary.column_dimensions[get_column_letter(col)].width=16

    chart_sheet['A1']='กราฟรายงานจากข้อมูลจริง'
    chart_sheet['A1'].font=Font(bold=True, size=16)
    chart_sheet['A3']='ชั้น'; chart_sheet['B3']='ฟันผุ'; chart_sheet['C3']='เหงือกอักเสบ'; chart_sheet['D3']='เร่งด่วน'
    for cell in chart_sheet[3]: cell.fill=header_fill; cell.font=bold; cell.border=border
    rr=4
    for g,v in by_grade.items():
        chart_sheet.cell(rr,1,g); chart_sheet.cell(rr,2,v['decay']); chart_sheet.cell(rr,3,v['gum']); chart_sheet.cell(rr,4,v['urgent'])
        for col in range(1,5): chart_sheet.cell(rr,col).border=border
        rr+=1
    bar=BarChart()
    bar.title='ฟันผุ / เหงือกอักเสบ / เร่งด่วน แยกชั้น'
    bar.y_axis.title='จำนวนคน'; bar.x_axis.title='ชั้นเรียน'
    data=Reference(chart_sheet,min_col=2,max_col=4,min_row=3,max_row=rr-1)
    cats=Reference(chart_sheet,min_col=1,min_row=4,max_row=rr-1)
    bar.add_data(data,titles_from_data=True); bar.set_categories(cats); bar.height=9; bar.width=20
    chart_sheet.add_chart(bar,'F3')

    pie_start=rr+3
    chart_sheet.cell(pie_start,1,'ภาวะโภชนาการ'); chart_sheet.cell(pie_start,2,'จำนวน')
    chart_sheet.cell(pie_start,1).fill=header_fill; chart_sheet.cell(pie_start,2).fill=header_fill
    pr=pie_start+1
    for label,count in nutrition.items():
        chart_sheet.cell(pr,1,label); chart_sheet.cell(pr,2,count); pr+=1
    pie=PieChart()
    pie.title='สัดส่วนภาวะโภชนาการ'
    labels=Reference(chart_sheet,min_col=1,min_row=pie_start+1,max_row=pr-1)
    data=Reference(chart_sheet,min_col=2,min_row=pie_start,max_row=pr-1)
    pie.add_data(data,titles_from_data=True); pie.set_categories(labels); pie.height=9; pie.width=12
    chart_sheet.add_chart(pie,'F22')
    for col in range(1,5): chart_sheet.column_dimensions[get_column_letter(col)].width=18

    bio=BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,download_name='dental_health_dashboard_with_charts.xlsx',as_attachment=True)

@app.route('/users',methods=['GET','POST'])
@login_required
@admin_required
def users():
    con=db()
    if request.method=='POST':
        con.execute('INSERT INTO users(username,password_hash,role,created_at) VALUES(?,?,?,?)',(request.form['username'],generate_password_hash(request.form['password']),request.form['role'],now())); con.commit()
    rows=con.execute('SELECT * FROM users ORDER BY id').fetchall(); con.close(); return render_template('users.html',users=rows)
@app.route('/user/<int:uid>/delete',methods=['POST'])
@login_required
@admin_required
def del_user(uid):
    if uid==session.get('user_id'): flash('ลบตัวเองไม่ได้','danger'); return redirect(url_for('users'))
    con=db(); con.execute('DELETE FROM users WHERE id=?',(uid,)); con.commit(); con.close(); return redirect(url_for('users'))

# Railway/gunicorn imports app.py directly, so initialize tables at import time.
init_db()

if __name__=='__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT',5001)), debug=True)
